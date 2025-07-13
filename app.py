from flask import Flask, request, redirect, render_template
from openpyxl import Workbook, load_workbook
import os

app = Flask(__name__)

@app.route('/index.html')
def home():
    return render_template('index.html')

@app.route('/contact.html')
def contact():
    return render_template('contact.html')

@app.route('/solutions.html')
def solutions():
    return render_template('solutions.html')

@app.route('/team.html')
def team():
    return render_template('team.html')

@app.route('/submit', methods=['POST'])
def submit():
    name = request.form['name']
    email = request.form['email']
    phone = request.form['phone']
    solution = request.form['solution']
    message = request.form['message']

    filename = 'contact_data.xlsx'
    if os.path.exists(filename):
        wb = load_workbook(filename)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(['Name', 'Email', 'Phone', 'Solution Required', 'Message'])

    ws.append([name, email, phone, solution, message])
    wb.save(filename)
    return redirect('index.html')

@app.route('/static/<path:filename>')
def static_files(filename):
    return app.send_static_file(filename)

if __name__ == '__main__':
    app.run(debug=True)
